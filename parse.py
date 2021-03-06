from bs4 import BeautifulSoup
import requests
import json
import csv
import openpyxl
import datetime
import time
import os
import random


def create_path():
    """Создает папку, потом забирает адрес и сохраняет туда файлы"""
    creation_time = str(datetime.datetime.now())[:10]
    basic_path = 'C:\\ParseProject'
    filepath = os.path.join(basic_path, creation_time)
    if not os.path.exists(filepath):
        os.makedirs(filepath)
    return filepath


class ExcelFile:
    """
    Преобразование результатов парсинга страниц в таблицу эксель;
    На вход принимаются лист - имена файлов(без расширения)
    название группы - будет названием файла и названием листа книги эксель
    и объект класса Html, просто созданный через инит Html
    Код запускается сразу при создании экземпляра
    """

    def __init__(self, lst, groupname, parsed, file_or_web=1, web=None):
        self.parsed = parsed
        self.lst = lst
        self.groupname = groupname

        print('Start to make .xlsx file')
        file = openpyxl.Workbook()
        """Создание формы таблицы
        """
        sheetsource = file.active
        sheetsource.title = self.groupname
        num = 2  # номер строки с которой начинается заполнение данными
        if file_or_web:
            for i in range(len(lst)):
                filename = str(lst[i]) + '.html'
                self.parsed.get_from_file(filename)  # метод открывает файл
                newdict = self.parsed.parse_card()  # метод парсит файл
                pagename = list(newdict.keys())[0]  # имя модели - ключ главного словаря
                dictkeys = list(newdict[pagename].keys())  # характеристики - ключи словаря
                head = self.table_making(dictkeys, newdict, pagename, 1)
                fill = self.table_making(dictkeys, newdict, pagename, 0)
                sheetsource.cell(row=1, column=1).value = 'Model name'
                for k in range(len(head)):
                    if sheetsource.cell(row=1, column=1).value is None:
                        sheetsource.cell(row=1, column=k + 2).value = head[k]
                    else:
                        sheetsource.cell(row=1, column=k + 2).value = head[k]
                    sheetsource.cell(row=num + i, column=1).value = pagename
                    sheetsource.cell(row=num + i, column=k + 2).value = fill[k]
        else:
            for i in range(len(lst)):

                self.parsed.get_content(web, {'pid': lst[i]})  # метод открывает страницу
                newdict = self.parsed.parse_card()  # метод парсит файл
                pagename = list(newdict.keys())[0]  # имя модели - ключ главного словаря
                dictkeys = list(newdict[pagename].keys())  # характеристики - ключи словаря
                head = self.table_making(dictkeys, newdict, pagename, 1)
                fill = self.table_making(dictkeys, newdict, pagename, 0)
                sheetsource.cell(row=1, column=1).value = 'Model name'
                for k in range(len(head)):
                    if sheetsource.cell(row=1, column=1).value is None:
                        sheetsource.cell(row=1, column=k + 2).value = head[k]
                    else:
                        sheetsource.cell(row=1, column=k + 2).value = head[k]
                    sheetsource.cell(row=num + i, column=1).value = pagename
                    sheetsource.cell(row=num + i, column=k + 2).value = fill[k]
        # TODO сделать одну функцию на этот блок чтобы не дублировать
        file.save(os.path.join(create_path(), groupname + '.xlsx'))
        print('Creating .xlsx is finished, filename ' + groupname)

    @staticmethod
    def table_making(keys: list, dictionary: dict, pagename: str, option=1):
        """
        Функция возвращает список - шапку - ключи, и список - значения ключей;
        Принимает ключи - список из ключей
        Словарь - словарь, получившийся после парсинга из которого будет создаваться заполнение по ключам;
        Название страницы - для заполнения значений ключей
        Параметр option - если 1 (по умолч., то создается шапка, если 0, то заполнение
        """
        if option:
            columns = []
            for i in range(len(keys)):
                columns.append(keys[i])
            return columns
        else:
            meaning = []
            for i in range(len(keys)):
                meaning.append(dictionary[pagename][keys[i]])
            return meaning


class WrongFile(BaseException):
    pass


class Html:
    """Создается экземпляр класса, далее пользуемся методами"""

    def __init__(self, prefix=None):
        self.prefix = prefix
        self.url = None
        self.params = None
        self.content = None

    def get_content(self, url, params=None):
        headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:71.0)'
                                 ' Gecko/20100101 Firefox/71.0', 'accept': '*/*'}
        self.url = url
        self.params = params
        response = requests.get(self.url, headers=headers, params=self.params, timeout=5)
        time.sleep(random.randint(1, 45))
        if response.status_code == 200:
            self.content = response.text

            print('Go to ', self.url, 'Status OK')
            return self.content
        else:
            print('Error')

    def get_from_file(self, filename):
        try:
            with open(filename, 'r') as file:
                self.content = file.read()
        except FileNotFoundError as err:
            print('Неправильное имя файла ', err)
            raise WrongFile


class AshParse(Html):
    def __init__(self, prefix=None):
        super().__init__(self)
        self.newdict = None
        self.pagename = None
        self.prefix = prefix

    def get_numbers(self):
        """ Собирает номера продуктов со страницы со списком
        """
        print('Start parsing ')
        soup = BeautifulSoup(self.content, 'html.parser')
        data = soup.find_all('div', class_="omega thirteen columns")
        numberlist = []

        for item in data:
            url = item.find('a').get('href')
            num = url[-3:]
            numberlist.append(int(num))
        return numberlist

    def title_page(self):
        """Собирает названия страницы"""
        print('Get page title for filename')
        soup = BeautifulSoup(self.content, 'html.parser')
        pagename = soup.find('h1', class_='product-title').get_text().replace('/', '')
        return pagename

    def parse(self):
        """Start to parse html file, returning dictionary with elements"""
        print('Start parsing ')
        soup = BeautifulSoup(self.content, 'html.parser')
        data = soup.find_all('div', class_="omega thirteen columns")
        self.pagename = soup.find('h1', class_='product-title').get_text().replace('/', '')
        picts = soup.find_all('div', class_="alpha three columns textcenter")
        # print(picts)
        self.newdict = {}
        # print(data)
        pictlst = []
        for item in picts:
            pict = item.find('img').get('src')
            pictlst.append(pict)
        k = 0  # переменная индекс списка piclst
        for item in data:
            title = item.find('p').get_text()
            url = item.find('a').get('href')
            lst = []
            acchead = item.find('thead').find_next('tr').find_next('th').get_text()
            stylehead = item.find('thead').find_next('tr').find_next('th'). \
                find_next('th').get_text()
            rangehead = item.find('thead').find_next('tr').find_next('th'). \
                find_next('th').find_next('th').get_text()
            accbody = item.find('tbody').find_next('td').get_text()
            stylebody = item.find('tbody').find_next('td').find_next('td').get_text()
            rangebody = item.find('tbody').find_next('td').find_next('td'). \
                find_next('td').get_text()
            lst.append(
                {acchead: accbody,
                 stylehead: stylebody,
                 rangehead: rangebody,
                 'url': url,
                 'picturl': 'https://www.ashcroft.eu' + pictlst[k]})
            k += 1  # прибавляем чтобы идти по списке
            self.newdict[title] = lst
        print('Parsing finished' + self.pagename)

        return self.newdict

    def parse_card(self):

        print('Start parsing cards')
        soup = BeautifulSoup(self.content, 'html.parser')
        self.pagename = soup.find('div', class_="center-column-content product-detail").find_next('h1').text
        basic = soup.find('div', id="overview-tab-content")
        "Блок полей для парсинга"
        bigphotos = soup.find('ul', class_='slides').find_next('a').get('href')
        smallphotos = soup.find('ul', class_='slides').find_next('li').get('data-thumb')
        description = basic.find_next('p').text
        features = basic.find_next('h5').text
        features_desr = basic.find_next('ul')
        featureslst = []
        for item in features_desr:
            featureslst.append(item.text)
        featuresstr = '; '.join(featureslst)
        spec = basic.find_next('h5').find_next('h5').text
        spec_descr = basic.find_next('ul').find_next('ul')
        speclst = []
        for i in spec_descr:
            speclst.append(i.text)
        specstr = '; '.join(speclst)
        appl = soup.find('div', id="applications-tab-content"). \
            text.replace('/', '-').replace('\\', '-')
        """ Этот блок приводит в нормальный вид описание сфер применения"""
        begin = 0
        fields = ''

        for i in range(len(appl)):
            if appl[i].islower() and appl[i + 1].isupper():
                fields += appl[begin + 1:i + 1] + ', '
                begin = i
        fields += appl[begin + 1:]
        datasheet = soup.find('div', class_='product-details widget'). \
            find_next('a', class_='detail-button').get('href')
        spaces = datasheet.count(' ')
        cleands = datasheet
        if spaces > 0:  # меняем пробелы на %20 чтобы файл скачивался
            for i in range(spaces):
                cleands = datasheet.replace(' ', '%20')

        self.newdict = {self.pagename: {
            'Description': description,
            features: featuresstr,
            spec: specstr,
            'Application': fields.strip(),
            'Large_picture': self.prefix[:-3] + bigphotos,
            'Small_picture': self.prefix[:-3] + smallphotos,
            'Datasheet': self.prefix[:-3] + cleands
        }
        }
        print('Parsing card ' + self.pagename + ' completed')
        return self.newdict

    def to_json(self):
        """ Method for the saving json file with the results dict"""
        print('Создаем .json файл')
        with open(self.pagename + '.json', 'w', encoding='utf8') as write_file:
            json.dump(self.newdict, write_file, indent=2)
        print('Создание файла json выполнено')

    def to_csv_product(self):
        """Create csv file"""
        print('Start to create .csv file')
        dictkeys = list(self.newdict[self.pagename].keys())

        with open(self.pagename + '.csv', 'w', newline='', encoding='utf8') as csvfile:
            writer = csv.writer(csvfile, delimiter=',')
            # writer.writeheader()
            writer.writerow(self.table_columns(dictkeys))

            writer.writerow(self.table_meaning(dictkeys))
        print('Creating .csv file is finished')

    def table_columns(self, lst):
        """ Функция делает шапку таблицы"""
        columns = []
        meaning = []
        for i in range(len(lst)):
            columns.append(lst[i])
            meaning.append(self.newdict[self.pagename][lst[i]])
        return columns

    def table_meaning(self, lst):
        """Функция заполняет значения таблицы под шапкой"""
        columns = []
        meaning = []
        for i in range(len(lst)):
            columns.append(lst[i])
            meaning.append(self.newdict[self.pagename][lst[i]])
        return meaning

    def excel_list_of_products(self, groupname: str, gidnumbers: list):
        """Принимает имя группы и список номеров -gid для файлов
         Парсинг из скачанных страниц, без get
         Складывает все значения в одну табличку"""
        print('Start to make .xlsx file')

        file = openpyxl.Workbook()
        """Создание формы нового файла"""
        sheetsource = file.active
        sheetsource.title = groupname
        num = 2
        for i in range(len(gidnumbers)):
            filename = str(gidnumbers[i]) + '.html'
            self.get_from_file(filename)
            self.parse_card()
            dictkeys = list(self.newdict[self.pagename].keys())
            head = self.table_columns(dictkeys)
            fill = self.table_meaning(dictkeys)
            sheetsource.cell(row=1, column=1).value = 'Model name'
            for k in range(len(head)):
                if sheetsource.cell(row=1, column=1).value is None:
                    sheetsource.cell(row=1, column=k + 2).value = head[k]
                else:
                    sheetsource.cell(row=1, column=k + 2).value = head[k]
                sheetsource.cell(row=num + i, column=1).value = self.pagename
                sheetsource.cell(row=num + i, column=k + 2).value = fill[k]

        file.save(groupname + '.xlsx')
        print('Creating .xlsx is finished, filename ' + groupname)

    def to_csv_line(self):
        """Create csv file"""
        print('Start to create .csv file')
        with open(self.pagename + '.csv', 'w', newline='', encoding='utf8') as csvfile:
            columns = ['Model', 'Accuracy', 'Dial Size and Style', 'Range', 'Url', 'Picture']
            writer = csv.writer(csvfile, delimiter=',')

            writer.writerow(columns)

            for key in self.newdict:
                writer.writerow(
                    [key,
                     self.newdict[key][0][list(self.newdict[key][0].keys())[0]],
                     self.newdict[key][0][list(self.newdict[key][0].keys())[1]],
                     self.newdict[key][0][list(self.newdict[key][0].keys())[2]],
                     self.newdict[key][0][list(self.newdict[key][0].keys())[3]],
                     self.newdict[key][0][list(self.newdict[key][0].keys())[4]],

                     ]
                )
        print('Creating .csv file is finished')

    def to_excel_line(self):
        """Create excel file"""
        print('Start to make .xlsx file')
        file = openpyxl.Workbook()
        """Создание формы нового файла"""
        sheetsource = file.active
        sheetsource.title = self.pagename
        columns = ['Model', 'Accuracy', 'Dial Size and Style', 'Range', 'Url', 'Picture']
        sheetsource['A1'] = columns[0]
        sheetsource['B1'] = columns[1]
        sheetsource['C1'] = columns[2]
        sheetsource['D1'] = columns[3]
        sheetsource['E1'] = columns[4]
        sheetsource['F1'] = columns[4]
        k = 2
        columnslst = ['A', 'B', 'C', 'D', 'E', 'F']
        for key in self.newdict:
            print(key)
            sheetsource[columnslst[0] + str(k)] = key.strip()
            sheetsource[columnslst[1] + str(k)] = self.newdict[key][0][list(self.newdict[key][0].keys())[0]]
            sheetsource[columnslst[2] + str(k)] = self.newdict[key][0][list(self.newdict[key][0].keys())[1]]
            sheetsource[columnslst[3] + str(k)] = self.newdict[key][0][list(self.newdict[key][0].keys())[2]]
            sheetsource[columnslst[4] + str(k)] = self.newdict[key][0][list(self.newdict[key][0].keys())[3]]
            sheetsource[columnslst[5] + str(k)] = self.newdict[key][0][list(self.newdict[key][0].keys())[4]]
            k += 1
        file.save(self.pagename + '.xlsx')
        print('Creating .xlsx is finished, filename ' + self.pagename)


class NksParse(Html):
    def __init__(self, prefix):
        super().__init__(prefix)

    def get_pages(self):  # TODO потом сделать приватным, или защищенным
        """Собирает страницы с линейкой продуктов, возвращает словарь
        :return dict
        """
        print('Собираем категории продуктов')
        soup = BeautifulSoup(self.content, 'html.parser')
        data = soup.find('div', class_="productTab-main")
        pages = {}
        subdata = soup.find_all('div', class_='productTab-sub')

        page = data.find_all('a')

        subpageslist = []
        for elem in subdata:

            subpages = {}
            subpage = elem.find_all('a')
            print(subpage)
            for i in subpage:
                title_page = i.get_text()
                result = i.get('href')
                subpages[title_page] = result
            subpageslist.append(subpages)
        print(subpageslist)
        koeff = 0
        for i in page:
            title_page = i.get_text()
            result = i.get('href')
            pages[title_page] = subpageslist[koeff]
            koeff += 1

        return pages

    def get_cards(self, jsonfile):
        """парсит каждую страничку, собирает номера карточек, возвращает список этих карточек
        и название категории
        :return dict
        """
        print('Получаем список карточек ')
        # pages = self.get_pages() - если парсить по по новой
        pages = self.from_json(jsonfile)
        page_dict = {}
        for key in pages:  # Страница со всеми продуктами
            cards = []  # Список из словарей - название категории - ссылки
            for elem in pages[key]:  # для каждого раздела
                print('Читаем страницу ' + key, ' раздел ', elem)
                card_url = self.prefix + pages[key][elem]
                content = self.get_content(card_url)
                card_dict = {}
                urls_lst = []
                pages_quantity = self._pages_qnt(content)
                if pages_quantity:  # если несколько страниц
                    print('Несколько вкладок')
                    for page in self._pages_qnt(content):
                        page_url = self.prefix + page
                        content = self.get_content(page_url)
                        soup = BeautifulSoup(content, 'html.parser')
                        data = soup.find_all('a', class_='productItem')
                        for item in data:
                            urls_lst.append(item.get('href'))
                    card_dict[elem] = urls_lst
                    cards.append(card_dict)
                else:
                    print('Одна вкладка')
                    soup = BeautifulSoup(content, 'html.parser')
                    data = soup.find_all('a', class_='productItem')
                    for item in data:
                        urls_lst.append(item.get('href'))
                    card_dict[elem] = urls_lst
                    cards.append(card_dict)
                page_dict[key] = cards
        return page_dict

    @staticmethod
    def to_json(name, dicti):
        """ Method for the saving json file with the results dict"""
        print('Создаем .json файл')
        with open(name + '.json', 'w', encoding='utf8') as write_file:
            json.dump(dicti, write_file, indent=2)
        print('Создание файла json выполнено')

    @staticmethod
    def _pages_qnt(content):
        """ количество вкладок на странице"""
        print('Получаем количество вкладок')
        soup = BeautifulSoup(content, 'html.parser')
        data = soup.find('div', class_='pager mt-20')

        if data:
            qnt = data.find_all('a')  # количество страниц
            pageslist = []
            for item in qnt:
                pageslist.append(item.get('href'))
            return pageslist
        else:
            return None

    @staticmethod
    def internal_pages(content):
        """
        На каждой странице еще есть вкладки - проверить # TODO
        :return: dict
        """
        print('Получаем подкатегории продуктов')
        int_pages_list = {}
        soup = BeautifulSoup(content, 'html.parser')
        data = soup.find('div', class_='productTab-sub')
        int_pages = data.find('ul').find_next('a', class_='is-current')

        title = int_pages.get_text()
        print(title)
        url = int_pages.get('href')
        int_pages_list[title] = url

        print(int_pages_list)
        return int_pages_list

    @staticmethod
    def from_json(file):
        with open(file, 'r', encoding='utf8') as f:
            cards_dict = json.load(f)
        return cards_dict

    def parse_cards(self, json_file):
        cards = self.from_json(json_file)
        full_dict = {}
        for key in cards:
            category_dict = {}
            for element in cards[key]:
                cards_dict = {}

                for cat in element:
                    print(cat)
                    for link in element[cat]:
                        print(link)
                        card_url = self.prefix + link
                        print(card_url)
                        content = self.get_content(card_url)
                        soup = BeautifulSoup(content, 'html.parser')
                        data = soup.find('div', class_='proDetail')
                        name = data.find('span', class_='proDetail-number').text
                        short = data.find('span', class_='proDetail-name').text
                        description = data.find('div', class_='proDetail-description').text
                        feature = data.find('ul', class_='proDetail-feature-ul')
                        if feature:
                            feature = feature.text
                        else:
                            feature = 'No information'

                        datasheet = data.find('a', class_='proDetail-download-btn')
                        if datasheet:
                            datasheet = self.prefix + datasheet.get('href')
                        else:
                            datasheet = 'No information'
                        image = data.find('img', class_='proDetail-img').get('src')
                        cards_dict[name] = {
                            'Short Description': short,
                            'Description': description,
                            'Features': feature,
                            'Datasheet': datasheet,
                            'Image': self.prefix + image,
                            'URL': card_url
                            }
                    category_dict[cat] = cards_dict

            full_dict[key] = category_dict
        return full_dict

#TODO сделать новый класс для создания эксель файла из jsona
